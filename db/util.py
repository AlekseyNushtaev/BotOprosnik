import datetime

import openpyxl
from sqlalchemy import select, insert, update

from db.models import Session, User


def add_user_to_db(id, username, first_name, last_name, time_start):
    with Session() as session:
        try:
            query = select(User).where(User.id == id)
            results = session.execute(query)
            if not results.all():
                stmt = insert(User).values(
                    id=id,
                    username=username,
                    first_name=first_name,
                    last_name=last_name,
                    time_start=time_start
                )
                session.execute(stmt)
                session.commit()
        except Exception as e:
            print(e)


def update_user_credit(id, credit):
    with Session() as session:
        try:
            stmt = update(User).where(User.id == id).values(credit=credit, time_credit=datetime.datetime.now())
            session.execute(stmt)
            session.commit()
        except Exception as e:
            print(e)


def update_user_ipoteka(id, ipoteka):
    with Session() as session:
        try:
            stmt = update(User).where(User.id == id).values(ipoteka=ipoteka, time_ipoteka=datetime.datetime.now())
            session.execute(stmt)
            session.commit()
        except Exception as e:
            print(e)


def update_user_house(id, house):
    with Session() as session:
        try:
            stmt = update(User).where(User.id == id).values(house=house, time_house=datetime.datetime.now())
            session.execute(stmt)
            session.commit()
        except Exception as e:
            print(e)


def update_user_auto(id, auto):
    with Session() as session:
        try:
            stmt = update(User).where(User.id == id).values(auto=auto, time_auto=datetime.datetime.now())
            session.execute(stmt)
            session.commit()
        except Exception as e:
            print(e)


def update_user_sdelki(id, sdelki):
    with Session() as session:
        try:
            stmt = update(User).where(User.id == id).values(sdelki=sdelki, time_sdelki=datetime.datetime.now())
            session.execute(stmt)
            session.commit()
        except Exception as e:
            print(e)


def update_user_pay_credit(id, pay_credit):
    with Session() as session:
        try:
            stmt = update(User).where(User.id == id).values(pay_credit=pay_credit, time_pay_credit=datetime.datetime.now())
            session.execute(stmt)
            session.commit()
        except Exception as e:
            print(e)


def update_user_debit(id, debit):
    with Session() as session:
        try:
            stmt = update(User).where(User.id == id).values(debit=debit, time_debit=datetime.datetime.now())
            session.execute(stmt)
            session.commit()
        except Exception as e:
            print(e)


def update_user_phone(id, phone):
    with Session() as session:
        try:
            stmt = update(User).where(User.id == id).values(phone=phone, time_phone=datetime.datetime.now())
            session.execute(stmt)
            session.commit()
        except Exception as e:
            print(e)


def update_user_blocked(id):
    with Session() as session:
        try:
            stmt = update(User).where(User.id == id).values(is_block=True)
            session.execute(stmt)
            session.commit()
        except Exception as e:
            print(e)


def update_user_unblocked(id):
    with Session() as session:
        try:
            stmt = update(User).where(User.id == id).values(is_block=False)
            session.execute(stmt)
            session.commit()
        except Exception as e:
            print(e)


def get_all_users():
    with Session() as session:
        query = select(User)
        users = session.execute(query)
        result = [['id', 'username', 'first_name', 'last_name', 'Время входа в бота',
                      'Сумма кредитов', 'Ипотека', 'Недвижимость', 'Автомобиль',
                   'Сделки за 3 года', 'Ежемесячный кредит', 'Доход', 'Телефон',
                   'Время оформления заявки', 'Сообщения', 'Блокировка бота']]
        for user in users.scalars():
            start = ''
            join = ''
            if user.time_start:
                start = user.time_start.strftime('%Y-%m-%d   %H:%M:%S')
            if user.time_phone:
                join = user.time_phone.strftime('%Y-%m-%d   %H:%M:%S')
            result.append([user.id, user.username, user.first_name, user.last_name,
                           start, user.credit, user.ipoteka, user.house, user.auto,
                           user.sdelki, user.pay_credit, user.debit, user.phone, join,
                           user.messages,user.is_block])
    return result


def get_all_users_unblock():
    with Session() as session:
        query = select(User).where(User.is_block == False)
        users = session.execute(query)
        result = []
        for user in users.scalars():
            result.append(user.id)
    return result


def excel_to_db(file):
    wb = openpyxl.load_workbook(file)
    sh = wb['Sheet1']
    for i in range(2, 61):
        with Session() as session:
            query = select(User).where(User.id == int(sh.cell(i, 1).value))
            results = session.execute(query)
        if not results.all():
            user_id = int(sh.cell(i, 1).value)
            username = sh.cell(i, 2).value
            first_name = sh.cell(i, 3).value
            last_name = sh.cell(i, 4).value
            time_start = sh.cell(i, 5).value
            credit = sh.cell(i, 6).value
            ipoteka = sh.cell(i, 7).value
            house = sh.cell(i, 8).value
            auto = sh.cell(i, 9).value
            sdelki = sh.cell(i, 10).value
            pay_credit = sh.cell(i, 11).value
            debit = sh.cell(i, 12).value
            phone = sh.cell(i, 13).value
            is_block = sh.cell(i, 15).value
            with Session() as session:
                stmt = insert(User).values(
                    id=user_id,
                    username=username,
                    first_name=first_name,
                    last_name=last_name,
                    time_start=time_start,
                    credit=credit,
                    ipoteka=ipoteka,
                    house=house,
                    auto=auto,
                    sdelki=sdelki,
                    pay_credit=pay_credit,
                    debit=debit,
                    phone=phone,
                    is_block=is_block
                )
                session.execute(stmt)
                session.commit()


def update_messages(to_user_id, user_id, username, text):
    with Session() as session:
        try:
            query = select(User).where(User.id == to_user_id)
            old_msg = session.execute(query).scalars().first().messages
            time_now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            new_msg = old_msg + f'{user_id}-{username}-({time_now}) - {text}\n'
            stmt = update(User).where(User.id == to_user_id).values(messages=new_msg)
            session.execute(stmt)
            session.commit()
        except Exception as e:
            print(e)




